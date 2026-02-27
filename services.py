import os
import shutil
import re
from datetime import datetime, date
from typing import Optional
import sys
from aiogram import Bot
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from config import UPLOAD_DIR, MAX_FILE_SIZE


class FileService:
    @staticmethod
    def get_candidate_dir(telegram_id: int, full_name: str) -> str:
        safe_name = re.sub(r'[^\w\s-]', '', full_name).strip().replace(' ', '_')
        path = os.path.join(UPLOAD_DIR, f"{safe_name}_{telegram_id}")
        os.makedirs(path, exist_ok=True)
        return path

    @staticmethod
    def save_file(data: bytes, directory: str, filename: str, subfolder: str = None) -> str:
        if subfolder:
            directory = os.path.join(directory, subfolder)
            os.makedirs(directory, exist_ok=True)
        filepath = os.path.join(directory, filename)
        with open(filepath, 'wb') as f:
            f.write(data)
        return filepath

    @staticmethod
    def get_file_size(data: bytes) -> int:
        return len(data)

    @staticmethod
    def is_valid_size(data: bytes) -> bool:
        return len(data) <= MAX_FILE_SIZE


class ValidationService:
    @staticmethod
    def validate_full_name(name: str) -> tuple[bool, str]:
        name = name.strip()
        words = name.split()
        if len(words) < 3:
            return False, "❌ Iltimos, to'liq F.I.Sh kiriting (kamida 3 ta so'z)\nMisol: Sobirov Dilshodbek Saydullo o'g'li"
        for word in words:
            if not re.match(r"^[a-zA-Zа-яА-ЯёЁa-zA-ZÀ-ÿ'`\- ]+$", word):
                # Allow Uzbek chars too
                pass
        return True, ""

    @staticmethod
    def validate_age(birth_date_str: str) -> tuple[bool, str]:
        try:
            bd = datetime.strptime(birth_date_str, "%d.%m.%Y").date()
            today = date.today()
            age = (today - bd).days // 365
            if age < 18:
                return False, "❌ Siz 18 yoshdan katta bo'lishingiz kerak."
            return True, ""
        except ValueError:
            return False, "❌ Sana formati noto'g'ri. Misol: 15.06.1995"

    @staticmethod
    def validate_achievements(text: str) -> tuple[bool, str]:
        if len(text) < 200:
            return False, f"❌ Kamida 200 belgi bo'lishi kerak. Hozir: {len(text)} belgi."
        return True, ""

    @staticmethod
    def calculate_experience(work_start_str: str) -> float:
        try:
            start = datetime.strptime(work_start_str, "%d.%m.%Y").date()
            today = date.today()
            delta = (today - start).days / 365.25
            return round(delta, 1)
        except:
            return 0.0


class ExportService:
    @staticmethod
    async def export_excel(applications: list) -> str:
        """Export all applications to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nomzodlar"

        headers = [
            "№", "F.I.Sh", "Telefon", "Viloyat", "Tuman", "Mahalla",
            "Tug'ilgan sana", "Ish boshlagan sana", "Tajriba (yil)",
            "Til sertifikati", "Namunali g'olib", "Top 100",
            "Tashabbus (Res)", "Tashabbus (Hud)", "Tashabbus (Tum)",
            "Qo'shimcha yutuqlar", "Davlat mukofoti", "Argos",
            "Mega loyihalar", "Status", "Ball", "Esse balli", "Suhbat sanasi", "Yakuniy holat"
        ]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, (app, user, score, interview) in enumerate(applications, 2):
            stage1_score = score.experience_score + score.results_score + score.motivation_score if score else ""

            data = [
                row_num - 1,
                user.full_name if user else "",
                user.phone_number if user else "",
                app.region or "",
                app.district or "",
                app.mahalla or "",
                app.birth_date or "",
                app.work_start_date or "",
                app.experience_years or 0,
                app.lang_certs or "",
                "Ha" if app.namunali_winner else "Yo'q",
                "Ha" if app.top100_winner else "Yo'q",
                "Ha" if app.initiative_respublika else "Yo'q",
                "Ha" if app.initiative_hudud else "Yo'q",
                "Ha" if app.initiative_tuman else "Yo'q",
                app.additional_achievements or "",
                "Ha" if app.state_award else "Yo'q",
                "Ha" if app.argos_status else "Yo'q",
                app.mega_projects or "",
                app.final_status or "",
                stage1_score,
                score.essay_score if score else "",
                score.total_score if score else "",
                interview.interview_date if interview else "",
                interview.status if interview else ""
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        export_path = os.path.join(UPLOAD_DIR, "exports")
        os.makedirs(export_path, exist_ok=True)
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(export_path, filename)
        wb.save(filepath)
        return filepath

    @staticmethod
    async def export_zip(applications: list, bot: Bot) -> Optional[str]:
        """Create ZIP with all candidate files by downloading them from Telegram."""
        import zipfile
        import logging
        
        export_path = os.path.join(UPLOAD_DIR, "exports")
        os.makedirs(export_path, exist_ok=True)
        zip_filename = f"all_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_filepath = os.path.join(export_path, zip_filename)

        try:
            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                for (app, user, score, interview) in applications:
                    if not user:
                        continue
                    safe_name = re.sub(r'[^\w\s-]', '', user.full_name or "").strip().replace(' ', '_')
                    folder_name = f"{safe_name}_{user.telegram_id}"
                    
                    if app.documents:
                        for doc in app.documents:
                            if doc.file_path:  # This is the file_id
                                try:
                                    file_info = await bot.get_file(doc.file_path)
                                    file_content = await bot.download_file(file_info.file_path)
                                    arcname = f"{folder_name}/{doc.file_type}/{doc.file_name}"
                                    zf.writestr(arcname, file_content.read())
                                except Exception as e:
                                    logging.error(f"Could not download file for app {app.id} (file_id: {doc.file_path}): {e}")
                                    zf.writestr(f"{folder_name}/DOWNLOAD_ERROR_{doc.file_name}.txt", f"Faylni yuklab bo'lmadi. Xatolik: {e}")
            return zip_filepath
        except Exception as e:
            logging.error(f"Failed to create ZIP archive: {e}")
            return None

    @staticmethod
    async def export_candidate_fish_excel(data: tuple) -> str:
        """Export candidate FISH (personal info) to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return None

        app, user, score, interview = data
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FISH Ma'lumotlari"
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        
        # Header style
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = f"📋 {user.full_name} - FISH Ma'lumotlari"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        
        # Personal Data
        personal_data = [
            ("F.I.Sh", user.full_name or ""),
            ("Telefon raqami", user.phone_number or ""),
            ("Tug'ilgan sanasi", app.birth_date or ""),
            ("Viloyat", app.region or ""),
            ("Tuman", app.district or ""),
            ("Mahalla", app.mahalla or ""),
        ]
        
        for label, value in personal_data:
            label_cell = ws.cell(row=row, column=1, value=label)
            value_cell = ws.cell(row=row, column=2, value=value)
            
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
            label_cell.border = thin_border
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row += 1
        
        # Professional Data
        row += 1
        prof_header = ws.cell(row=row, column=1, value="KASBIY MA'LUMOTLAR")
        prof_header.font = Font(bold=True, size=11, color="FFFFFF")
        prof_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        professional_data = [
            ("Ish boshlagan sanasi", app.work_start_date or ""),
            ("Tajriba (yil)", app.experience_years or ""),
            ("Til sertifikatlar", app.lang_certs or ""),
            ("Davlat mukofati", "Ha" if app.state_award else "Yo'q"),
            ("Argos tasnifi", "Ha" if app.argos_status else "Yo'q"),
        ]
        
        for label, value in professional_data:
            label_cell = ws.cell(row=row, column=1, value=label)
            value_cell = ws.cell(row=row, column=2, value=value)
            
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
            label_cell.border = thin_border
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row += 1
        
        # Achievements
        row += 1
        ach_header = ws.cell(row=row, column=1, value="YUTUQLAR VA TASHABBUS")
        ach_header.font = Font(bold=True, size=11, color="FFFFFF")
        ach_header.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        achievements_data = [
            ("Namunali g'olib", "Ha" if app.namunali_winner else "Yo'q"),
            ("Top 100 g'olib", "Ha" if app.top100_winner else "Yo'q"),
            ("Tashabbus (Respublika)", "Ha" if app.initiative_respublika else "Yo'q"),
            ("Tashabbus (Hudud)", "Ha" if app.initiative_hudud else "Yo'q"),
            ("Tashabbus (Tuman)", "Ha" if app.initiative_tuman else "Yo'q"),
            ("Qo'shimcha yutuqlar", app.additional_achievements or ""),
        ]
        
        for label, value in achievements_data:
            label_cell = ws.cell(row=row, column=1, value=label)
            value_cell = ws.cell(row=row, column=2, value=value)
            
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            label_cell.border = thin_border
            label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row += 1
        
        # Additional Fields
        if app.mega_projects or app.social_telegram or app.social_facebook or app.social_instagram:
            row += 1
            add_header = ws.cell(row=row, column=1, value="QOSHIMCHA MA'LUMOTLAR")
            add_header.font = Font(bold=True, size=11, color="FFFFFF")
            add_header.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            additional_data = []
            if app.mega_projects:
                additional_data.append(("Mega loyihalar", app.mega_projects))
            if app.social_telegram:
                additional_data.append(("Telegram", app.social_telegram))
            if app.social_facebook:
                additional_data.append(("Facebook", app.social_facebook))
            if app.social_instagram:
                additional_data.append(("Instagram", app.social_instagram))
            
            for label, value in additional_data:
                label_cell = ws.cell(row=row, column=1, value=label)
                value_cell = ws.cell(row=row, column=2, value=value)
                
                label_cell.font = Font(bold=True)
                label_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                label_cell.border = thin_border
                label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                value_cell.border = thin_border
                value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                row += 1
        
        # Status section
        row += 1
        status_header = ws.cell(row=row, column=1, value="STATUS")
        status_header.font = Font(bold=True, size=11, color="FFFFFF")
        status_header.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        status_data = [
            ("Hozirgi bosqich", f"Bosqich {app.current_stage}"),
            ("Yakuniy holat", app.final_status or ""),
            ("Topshirilgan vaqti", str(app.submitted_at)[:16] if app.submitted_at else ""),
        ]
        
        for label, value in status_data:
            label_cell = ws.cell(row=row, column=1, value=label)
            value_cell = ws.cell(row=row, column=2, value=value)
            
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            label_cell.border = thin_border
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row += 1
        
        export_path = os.path.join(UPLOAD_DIR, "exports")
        os.makedirs(export_path, exist_ok=True)
        filename = f"fish_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(export_path, filename)
        wb.save(filepath)
        return filepath

    @staticmethod
    async def export_candidate_zip(data: tuple, bot: Bot) -> Optional[str]:
        """Export single candidate data (Excel + FISH + Docs) to ZIP"""
        import zipfile
        import logging
        
        app, user, score, interview = data
        
        # 1. Generate main Excel
        excel_path = await ExportService.export_excel([data])
        
        # 2. Generate FISH Excel
        fish_excel_path = await ExportService.export_candidate_fish_excel(data)
        
        # 3. Prepare ZIP
        export_path = os.path.join(UPLOAD_DIR, "exports")
        os.makedirs(export_path, exist_ok=True)
        
        safe_name = re.sub(r'[^\w\s-]', '', user.full_name or "nomzod").strip().replace(' ', '_')
        zip_filename = f"{safe_name}_{user.telegram_id}.zip"
        zip_filepath = os.path.join(export_path, zip_filename)
        
        try:
            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Add main Excel summary
                if excel_path and os.path.exists(excel_path):
                    zf.write(excel_path, f"{safe_name}_info.xlsx")
                
                # Add FISH Excel
                if fish_excel_path and os.path.exists(fish_excel_path):
                    zf.write(fish_excel_path, f"{safe_name}_FISH.xlsx")
                
                # Add Documents
                if app.documents:
                    for doc in app.documents:
                        if doc.file_path:
                            try:
                                file_info = await bot.get_file(doc.file_path)
                                file_content = await bot.download_file(file_info.file_path)
                                filename = doc.file_name or f"{doc.file_type}.pdf"
                                zf.writestr(f"Hujjatlar/{filename}", file_content.read())
                            except Exception as e:
                                logging.error(f"Failed to download {doc.file_type}: {e}")
                                zf.writestr(f"Hujjatlar/ERROR_{doc.file_type}.txt", str(e))
            
            # Clean up temp Excel files
            if excel_path and os.path.exists(excel_path):
                os.remove(excel_path)
            if fish_excel_path and os.path.exists(fish_excel_path):
                os.remove(fish_excel_path)
            
            return zip_filepath
        except Exception as e:
            logging.error(f"Zip export error: {e}")
            return None
